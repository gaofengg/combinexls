# encoding: utf-8
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import time
from openpyxl.styles import Font
import argparse

from utils import xl_trans


def copy_cell_style(source_cell, target_cell):
    # 复制单元格样式
    target_cell._style = source_cell._style


def combinexsl():
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="combine xlsx workspace path")
    args = parser.parse_args()
    work_path = args.path

    gen_path = work_path + 'xlsx/'
    xl_trans.to_xlsx(work_path)
    # 使用 openpyxl 读取 ./xlsx/ 目录下的第一个文件
    file_list = [file for file in os.listdir(gen_path) if file.endswith(".xlsx")]
    workbook_first = Workbook()
    worksheet_first = workbook_first.create_sheet("Sheet1")
    for index, file in enumerate(file_list):
        # print(index, file)
        if index == 0:
            # 使用openpyxl读取第一个Excel文件
            workbook_first = load_workbook(gen_path + file)
            worksheet_first_name = workbook_first.sheetnames[0]
            worksheet_first = workbook_first[worksheet_first_name]
            continue
        workbook_next = load_workbook(gen_path + file)
        worksheet_next_name = workbook_next.sheetnames[0]
        worksheet_next = workbook_next[worksheet_next_name]

        # 一行一行写
        # for row in worksheet_next.iter_rows(min_row=3, values_only=True):
        #     worksheet_first.append(row)

        # 获取第一个工作表的最后一行的行号
        last_row_first = worksheet_first.max_row

        for row_index, row in enumerate(
            worksheet_next.iter_rows(min_row=3, values_only=False), start=1
        ):
            for col_index, cell in enumerate(row, start=1):
                # 将数据写入第一个工作表
                target_cell = worksheet_first.cell(
                    row=last_row_first + row_index, column=col_index
                )
                target_cell.value = cell.value

                # 复制样式
                copy_cell_style(cell, target_cell)

    combined_file_name = work_path + "combined_" + str(int(time.time())) + ".xlsx"
    workbook_first.save(combined_file_name)
    
    shutil.rmtree(gen_path[:-1])


if __name__ == "__main__":
    combinexsl()
